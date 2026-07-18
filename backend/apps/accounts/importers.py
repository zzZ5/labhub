from __future__ import annotations

import re
from typing import BinaryIO

from django.contrib.auth import get_user_model
from django.db import transaction

from apps.students.models import StudentProfile

from .models import Role, UserProfile
from .serializers import AdminUserCreateSerializer
from .services import SYSTEM_PERMISSION_CODES, SYSTEM_ROLES


User = get_user_model()

SCHOOL_IDENTITIES = {
    "硕博导师": UserProfile.SchoolIdentity.PI,
    "博士后": UserProfile.SchoolIdentity.POSTDOC,
    "博士生": UserProfile.SchoolIdentity.PHD,
    "硕士生": UserProfile.SchoolIdentity.MASTER,
    "本科生": UserProfile.SchoolIdentity.UNDERGRADUATE,
    "其他成员": UserProfile.SchoolIdentity.OTHER,
    "其他": UserProfile.SchoolIdentity.OTHER,
}
MEMBERSHIP_STATUSES = {
    "在组": UserProfile.MembershipStatus.ACTIVE,
    "已毕业/离组": UserProfile.MembershipStatus.FORMER,
    "已毕业": UserProfile.MembershipStatus.FORMER,
    "已离组": UserProfile.MembershipStatus.FORMER,
}
SYSTEM_PERMISSIONS = {
    "网站编辑": "editor",
    "资料管理员": "document_manager",
    "仪器管理员": "instrument_manager",
    "系统管理员": "admin",
}
STUDENT_DEGREES = {
    UserProfile.SchoolIdentity.UNDERGRADUATE: StudentProfile.DegreeType.UNDERGRADUATE,
    UserProfile.SchoolIdentity.MASTER: StudentProfile.DegreeType.MASTER,
    UserProfile.SchoolIdentity.PHD: StudentProfile.DegreeType.PHD,
}


def clean_cell(value) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    if text.endswith(".0") and text[:-2].isdigit():
        return text[:-2]
    return text


def normalize_identity(value: str) -> str:
    text = value.strip()
    if text in UserProfile.SchoolIdentity.values:
        return text
    return SCHOOL_IDENTITIES.get(text, "")


def normalize_membership_status(value: str) -> str:
    text = value.strip()
    if not text:
        return UserProfile.MembershipStatus.ACTIVE
    if text in UserProfile.MembershipStatus.values:
        return text
    return MEMBERSHIP_STATUSES.get(text, "")


def normalize_permissions(value: str) -> list[str]:
    values = [item.strip() for item in re.split(r"[,，、;；]+", value) if item.strip()]
    permissions = []
    for item in values:
        code = item if item in SYSTEM_PERMISSION_CODES else SYSTEM_PERMISSIONS.get(item, "")
        if code and code not in permissions:
            permissions.append(code)
    return permissions


def normalize_boolean(value: str, *, default: bool = False) -> bool:
    text = value.strip().lower()
    if not text:
        return default
    return text in {"1", "true", "yes", "y", "是", "生成", "已审核"}


def import_sheet(workbook):
    for sheet in workbook.worksheets:
        if "账号" in sheet.title or "成员" in sheet.title:
            return sheet
    return workbook.worksheets[-1]


def serializer_error(errors) -> str:
    for field, messages in errors.items():
        if isinstance(messages, dict):
            return serializer_error(messages)
        if isinstance(messages, (list, tuple)):
            message = str(messages[0]) if messages else "内容不正确"
        else:
            message = str(messages)
        labels = {
            "email": "邮箱",
            "username": "账号名",
            "password": "初始密码",
            "school_identity": "学校身份",
            "membership_status": "成员状态",
            "system_roles": "系统权限",
        }
        return f"{labels.get(field, field)}：{message}"
    return "账号信息不正确"


def ensure_system_roles():
    for code, name, description in SYSTEM_ROLES:
        Role.objects.update_or_create(
            code=code,
            defaults={"name": name, "description": description, "is_system": True},
        )


def import_accounts_from_excel(file_obj: BinaryIO, *, request) -> dict:
    from openpyxl import load_workbook

    workbook = load_workbook(file_obj, data_only=True)
    sheet = import_sheet(workbook)
    header_values = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if not header_values:
        raise ValueError("Excel 中没有账号数据。")
    headers = [clean_cell(value) for value in header_values]
    if "姓名" not in headers or "初始密码" not in headers:
        raise ValueError("账号表至少需要“姓名、初始密码”列。")

    ensure_system_roles()
    created = 0
    skipped = 0
    failed = 0
    student_profiles = 0
    total = 0
    issues = []

    for row_number, values in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        data = {header: clean_cell(value) for header, value in zip(headers, values) if header}
        if not any(data.values()):
            continue
        total += 1
        real_name = data.get("姓名", "")
        email = data.get("邮箱", "").lower()
        username = data.get("账号名", "") or email
        password = data.get("初始密码", "")
        identity = normalize_identity(data.get("学校身份", ""))
        membership_status = normalize_membership_status(data.get("成员状态", ""))
        permissions = normalize_permissions(data.get("系统权限", ""))

        if not real_name:
            failed += 1
            issues.append({"row": row_number, "status": "failed", "message": "姓名不能为空。"})
            continue
        if not email and not username:
            failed += 1
            issues.append({"row": row_number, "status": "failed", "message": "邮箱和账号名至少填写一个。"})
            continue
        if not identity:
            failed += 1
            issues.append({"row": row_number, "status": "failed", "message": "学校身份不正确。"})
            continue
        if not membership_status:
            failed += 1
            issues.append({"row": row_number, "status": "failed", "message": "成员状态不正确。"})
            continue
        raw_permissions = [item.strip() for item in re.split(r"[,，、;；]+", data.get("系统权限", "")) if item.strip()]
        if len(permissions) != len(dict.fromkeys(raw_permissions)):
            failed += 1
            issues.append({"row": row_number, "status": "failed", "message": "系统权限中包含无法识别的内容。"})
            continue

        duplicate = User.objects.filter(email__iexact=email).first() if email else None
        duplicate = duplicate or (User.objects.filter(username__iexact=username).first() if username else None)
        if duplicate:
            skipped += 1
            issues.append({"row": row_number, "status": "skipped", "message": f"账号已存在：{duplicate.email or duplicate.username}"})
            continue

        payload = {
            "real_name": real_name,
            "email": email,
            "username": username,
            "password": password,
            "school_identity": identity,
            "membership_status": membership_status,
            "is_approved": normalize_boolean(data.get("审核状态", ""), default=True),
            "system_roles": permissions,
        }
        try:
            with transaction.atomic():
                serializer = AdminUserCreateSerializer(data=payload, context={"request": request})
                if not serializer.is_valid():
                    raise ValueError(serializer_error(serializer.errors))
                user = serializer.save()
                should_create_profile = normalize_boolean(data.get("生成学生档案", ""))
                if should_create_profile:
                    degree = STUDENT_DEGREES.get(identity)
                    if not degree:
                        raise ValueError("只有本科生、硕士生或博士生可以生成学生档案。")
                    StudentProfile.objects.create(
                        user=user,
                        name=real_name,
                        degree_type=degree,
                        grade=data.get("年级", ""),
                    )
                    student_profiles += 1
            created += 1
        except ValueError as exc:
            failed += 1
            issues.append({"row": row_number, "status": "failed", "message": str(exc)})

    return {
        "total": total,
        "created": created,
        "skipped": skipped,
        "failed": failed,
        "student_profiles": student_profiles,
        "issues": issues,
    }
