from __future__ import annotations

import re
from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import BinaryIO

from django.core.files.base import ContentFile
from django.db import transaction
from django.utils.text import slugify

from apps.members.models import Member
from apps.news.models import Visibility
from apps.publications.models import Award, Patent, Project, Publication


IMPORT_KINDS = {"members", "publications", "projects", "patents", "awards"}


@dataclass(frozen=True)
class ImportImage:
    row_number: int
    filename: str
    content: bytes


def clean_cell(value) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    text = str(value).strip()
    if text.endswith(".0") and text[:-2].isdigit():
        return text[:-2]
    return text


def get_value(row: dict[str, str], *keys: str, default: str = "") -> str:
    for key in keys:
        value = row.get(key)
        if value not in (None, ""):
            return value.strip()
    return default


def parse_int(value: str, default: int = 0) -> int:
    if value in ("", None):
        return default
    try:
        return int(float(str(value).strip()))
    except (TypeError, ValueError):
        return default


def parse_decimal(value: str, default: Decimal | None = Decimal("0.00")) -> Decimal | None:
    if value in ("", None):
        return default
    try:
        return Decimal(str(value).strip())
    except (InvalidOperation, ValueError):
        return default


def parse_date(value: str):
    text = clean_cell(value)
    if not text:
        return None
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d", "%Y-%m", "%Y/%m", "%Y"):
        try:
            parsed = datetime.strptime(text, fmt)
            return parsed.date()
        except ValueError:
            continue
    return None


def parse_publication_citation(value: str) -> dict[str, str]:
    text = re.sub(r"\s+", " ", clean_cell(value)).strip()
    if not text:
        return {}
    text = re.sub(r"^\[\d+\]\s*", "", text)
    doi_match = re.search(r"(?i)\bdoi[:：]?\s*(10\.\S+)", text)
    doi = doi_match.group(1).rstrip(" .。") if doi_match else ""
    cleaned = re.sub(r"(?i)\bdoi[:：]?\s*10\.\S+", "", text).strip(" .。")

    year_match = re.search(r"\b(19\d{2}|20\d{2})[a-z]?\b", cleaned, flags=re.IGNORECASE)
    year = year_match.group(1) if year_match else ""
    authors = title = journal = journal_meta = ""
    if year_match:
        before_year = cleaned[: year_match.start()].strip(" ,，.。")
        after_year = cleaned[year_match.end():].strip(" ,，.。")
        year_is_after_authors = bool(re.search(r"\b(19\d{2}|20\d{2})[a-z]?\.\s+", cleaned, flags=re.IGNORECASE))
        if year_is_after_authors:
            authors = before_year
            title_split = re.split(r"\.\s+|。\s*", after_year, maxsplit=1)
            title = title_split[0].strip(" .。") if title_split else ""
            journal_meta = title_split[1] if len(title_split) == 2 else ""
            journal = journal_meta.strip(" .。")
        else:
            parts = [part.strip(" ,，.。") for part in re.split(r"\.\s+|。\s*", before_year) if part.strip(" ,，.。")]
            if len(parts) >= 3:
                authors = ". ".join(parts[:-2])
                title = parts[-2]
                journal = parts[-1]
            elif len(parts) == 2:
                authors, title = parts
            else:
                title = before_year
            journal_meta = f"{journal}, {after_year}".strip(" ,，")
    else:
        title = cleaned

    journal = (journal_meta or journal).strip(" .。")
    volume = issue = pages = ""
    meta_match = re.search(
        r"^(?P<journal>.+?)[\.,。]\s*(?:19\d{2}|20\d{2})?[a-z]?\s*[,，]?\s*(?P<volume>\d+)(?:\((?P<issue>[^)]+)\))?(?:\s*[:：,，]\s*|\s+)?(?P<pages>[A-Za-z]?\d+(?:[-–—]\d+)?|e\d+)?$",
        journal,
    )
    if meta_match:
        journal = meta_match.group("journal") or ""
        volume = meta_match.group("volume") or ""
        issue = meta_match.group("issue") or ""
        pages = (meta_match.group("pages") or "").replace("–", "-").replace("—", "-")

    return {
        "authors": authors.strip(" .。"),
        "title": title.strip(" .。"),
        "journal": journal.strip(" .。"),
        "year": year,
        "volume": volume,
        "issue": issue,
        "pages": pages,
        "doi": doi,
    }

def normalize_visibility(value: str) -> str:
    text = (value or "").strip().lower()
    if text in {"members", "member", "组内", "成员可见", "组内成员可见"}:
        return Visibility.MEMBERS
    if text in {"admins", "admin", "管理员", "管理员可见"}:
        return Visibility.ADMINS
    return Visibility.PUBLIC


def normalize_bool(value: str, default: bool = True) -> bool:
    text = (value or "").strip().lower()
    if not text:
        return default
    return text not in {"0", "false", "否", "不公开", "no", "n"}


def row_dict(headers: list[str], values) -> dict[str, str]:
    return {header: clean_cell(value) for header, value in zip(headers, values) if header}


def read_rows_from_excel(file_obj: BinaryIO) -> tuple[list[dict[str, str]], list[ImportImage]]:
    from openpyxl import load_workbook

    workbook = load_workbook(file_obj, data_only=True)
    if "导入数据" in workbook.sheetnames:
        sheet = workbook["导入数据"]
    elif len(workbook.worksheets) > 1:
        sheet = workbook.worksheets[1]
    else:
        sheet = workbook.worksheets[0]
    headers = [clean_cell(value) for value in next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))]
    rows = [row_dict(headers, values) for values in sheet.iter_rows(min_row=2, values_only=True)]
    rows = [row for row in rows if any(row.values())]

    images: list[ImportImage] = []
    for index, image in enumerate(getattr(sheet, "_images", []), start=1):
        marker = getattr(image.anchor, "_from", None)
        if marker is None:
            continue
        row_number = marker.row + 1
        ext = image_extension(image)
        images.append(ImportImage(row_number=row_number, filename=f"cms-import-{row_number}-{index}.{ext}", content=image._data()))
    return rows, images


def image_extension(image) -> str:
    fmt = (getattr(image, "format", "") or "").lower()
    if fmt in {"jpeg", "jpg"}:
        return "jpg"
    if fmt in {"png", "gif", "webp"}:
        return fmt
    suffix = Path(getattr(image, "path", "") or "").suffix.lower().lstrip(".")
    return suffix or "png"


def images_by_row(images: list[ImportImage]) -> dict[int, ImportImage]:
    grouped: dict[int, ImportImage] = {}
    for image in images:
        grouped.setdefault(image.row_number, image)
    return grouped


def upsert_record(model, lookup: dict, defaults: dict):
    instance = model.objects.filter(**lookup).order_by("id").first()
    if instance:
        for field, value in defaults.items():
            setattr(instance, field, value)
        instance.save()
        return instance, False
    return model.objects.create(**defaults), True


def import_rows(file_obj: BinaryIO, filename: str, kind: str) -> dict[str, int]:
    if kind not in IMPORT_KINDS:
        raise ValueError("不支持的导入类型。")

    lower_name = filename.lower()
    if not lower_name.endswith(".xlsx"):
        raise ValueError("请上传 .xlsx 文件。")
    rows, images = read_rows_from_excel(file_obj)

    image_lookup = images_by_row(images)
    handlers = {
        "members": import_members,
        "publications": import_publications,
        "projects": import_projects,
        "patents": import_patents,
        "awards": import_awards,
    }
    return handlers[kind](rows, image_lookup)


@transaction.atomic
def import_members(rows: list[dict[str, str]], image_lookup: dict[int, ImportImage]) -> dict[str, int]:
    created = updated = images = skipped = 0
    for index, row in enumerate(rows, start=2):
        name = get_value(row, "姓名", "name")
        if not name:
            skipped += 1
            continue
        email = get_value(row, "邮箱", "email")
        lookup = {"email__iexact": email} if email else {"name": name}
        defaults = {
            "name": name,
            "role_type": get_value(row, "身份头衔", "身份", "role_type"),
            "research_direction": get_value(row, "研究方向", "research_direction"),
            "email": email,
            "profile": get_value(row, "简介", "个人简介", "profile"),
            "join_date": parse_date(get_value(row, "加入日期", "join_date")),
            "graduation_date": parse_date(get_value(row, "毕业日期", "graduation_date")),
            "destination": get_value(row, "毕业去向", "去向", "destination"),
            "is_public": normalize_bool(get_value(row, "是否公开", "is_public"), default=True),
            "sort_order": parse_int(get_value(row, "展示排序", "排序", "sort_order")),
        }
        member, was_created = Member.objects.update_or_create(defaults=defaults, **lookup)
        created += int(was_created)
        updated += int(not was_created)
        image = image_lookup.get(index)
        if image:
            member.avatar.save(f"{slugify(name, allow_unicode=True) or 'member'}-{image.filename}", ContentFile(image.content), save=True)
            images += 1
    return {"created": created, "updated": updated, "skipped": skipped, "images": images, "total": len(rows)}


@transaction.atomic
def import_publications(rows: list[dict[str, str]], image_lookup: dict[int, ImportImage] | None = None) -> dict[str, int]:
    created = updated = skipped = 0
    for row in rows:
        parsed = parse_publication_citation(get_value(row, "GB/T 7714-2025格式引文", "国标格式", "参考文献", "引文", "citation", "citation_text"))
        title = get_value(row, "论文题目", "题目", "title", default=parsed.get("title", ""))
        if not title:
            skipped += 1
            continue
        doi = get_value(row, "DOI", "doi", default=parsed.get("doi", ""))
        year = parse_int(get_value(row, "年份", "year", default=parsed.get("year", "")), default=datetime.now().year)
        lookup = {"doi__iexact": doi} if doi else {"title": title, "year": year}
        defaults = {
            "title": title,
            "authors": get_value(row, "作者", "authors", default=parsed.get("authors", "")),
            "journal": get_value(row, "期刊", "journal", default=parsed.get("journal", "")),
            "year": year,
            "volume": get_value(row, "卷", "volume", default=parsed.get("volume", "")),
            "issue": get_value(row, "期", "issue", default=parsed.get("issue", "")),
            "pages": get_value(row, "页码", "pages", default=parsed.get("pages", "")),
            "doi": doi,
            "impact_factor": parse_decimal(get_value(row, "影响因子", "impact_factor"), default=None),
            "jcr_partition": get_value(row, "JCR分区", "JCR 分区", "jcr_partition"),
            "cas_partition": get_value(row, "中科院分区", "cas_partition"),
            "abstract": get_value(row, "摘要", "abstract"),
            "visibility": normalize_visibility(get_value(row, "可见范围", "visibility")),
            "sort_order": parse_int(get_value(row, "首页排序", "排序", "sort_order")),
        }
        _, was_created = upsert_record(Publication, lookup, defaults)
        created += int(was_created)
        updated += int(not was_created)
    return {"created": created, "updated": updated, "skipped": skipped, "images": 0, "total": len(rows)}


@transaction.atomic
def import_projects(rows: list[dict[str, str]], image_lookup: dict[int, ImportImage] | None = None) -> dict[str, int]:
    created = updated = skipped = 0
    for row in rows:
        title = get_value(row, "项目名称", "title")
        if not title:
            skipped += 1
            continue
        project_number = get_value(row, "项目编号", "project_number")
        lookup = {"project_number": project_number} if project_number else {"title": title}
        defaults = {
            "title": title,
            "project_number": project_number,
            "funding_source": get_value(row, "资助来源", "项目来源", "funding_source"),
            "principal_investigator": get_value(row, "负责人", "principal_investigator"),
            "start_date": parse_date(get_value(row, "开始日期", "start_date")),
            "end_date": parse_date(get_value(row, "结束日期", "end_date")),
            "amount": parse_decimal(get_value(row, "经费", "amount")),
            "status": get_value(row, "状态", "status"),
            "visibility": normalize_visibility(get_value(row, "可见范围", "visibility")),
            "description": get_value(row, "说明", "description"),
            "sort_order": parse_int(get_value(row, "首页排序", "排序", "sort_order")),
        }
        _, was_created = upsert_record(Project, lookup, defaults)
        created += int(was_created)
        updated += int(not was_created)
    return {"created": created, "updated": updated, "skipped": skipped, "images": 0, "total": len(rows)}


@transaction.atomic
def import_patents(rows: list[dict[str, str]], image_lookup: dict[int, ImportImage] | None = None) -> dict[str, int]:
    created = updated = skipped = 0
    for row in rows:
        title = get_value(row, "专利名称", "title")
        if not title:
            skipped += 1
            continue
        patent_number = get_value(row, "专利号", "patent_number")
        lookup = {"patent_number": patent_number} if patent_number else {"title": title}
        defaults = {
            "title": title,
            "patent_number": patent_number,
            "inventors": get_value(row, "发明人", "inventors"),
            "application_date": parse_date(get_value(row, "申请日期", "application_date")),
            "authorization_date": parse_date(get_value(row, "授权日期", "authorization_date")),
            "status": get_value(row, "状态", "status"),
            "visibility": normalize_visibility(get_value(row, "可见范围", "visibility")),
            "sort_order": parse_int(get_value(row, "首页排序", "排序", "sort_order")),
        }
        _, was_created = upsert_record(Patent, lookup, defaults)
        created += int(was_created)
        updated += int(not was_created)
    return {"created": created, "updated": updated, "skipped": skipped, "images": 0, "total": len(rows)}


@transaction.atomic
def import_awards(rows: list[dict[str, str]], image_lookup: dict[int, ImportImage]) -> dict[str, int]:
    created = updated = images = skipped = 0
    for index, row in enumerate(rows, start=2):
        title = get_value(row, "奖项名称", "获奖名称", "title")
        if not title:
            skipped += 1
            continue
        award_date = parse_date(get_value(row, "获奖日期", "award_date"))
        lookup = {"title": title, "award_date": award_date}
        defaults = {
            "title": title,
            "award_level": get_value(row, "奖项等级", "奖励等级", "award_level"),
            "award_date": award_date,
            "participants": get_value(row, "参与人员", "participants"),
            "description": get_value(row, "说明", "description"),
            "visibility": normalize_visibility(get_value(row, "可见范围", "visibility")),
            "sort_order": parse_int(get_value(row, "首页排序", "排序", "sort_order")),
        }
        award, was_created = upsert_record(Award, lookup, defaults)
        created += int(was_created)
        updated += int(not was_created)
        image = image_lookup.get(index)
        if image:
            award.image.save(f"{slugify(title, allow_unicode=True) or 'award'}-{image.filename}", ContentFile(image.content), save=True)
            images += 1
    return {"created": created, "updated": updated, "skipped": skipped, "images": images, "total": len(rows)}
