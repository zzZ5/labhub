from __future__ import annotations

from collections import defaultdict
from dataclasses import dataclass
from pathlib import Path
from typing import BinaryIO

from django.contrib.auth import get_user_model
from django.core.files.base import ContentFile
from django.db import transaction
from django.utils.text import slugify

from .models import Instrument, InstrumentCategory


DEFAULT_CATEGORY_NAME = "实验室设备"
DEFAULT_CATEGORY_SLUG = "lab-equipment"


@dataclass(frozen=True)
class InstrumentImportRow:
    index: int
    name: str
    model: str
    owner: str
    is_asset: str
    location_detail: str
    remark: str
    manager_name: str
    notes: str
    status: str
    occurrence: int


def clean_cell(value) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    if text.endswith(".0") and text[:-2].isdigit():
        return text[:-2]
    return text


def sheet_by_keywords(workbook, keywords: list[str]):
    for sheet_name in workbook.sheetnames:
        if any(keyword in sheet_name for keyword in keywords):
            return workbook[sheet_name]
    return None


def row_dict(headers: list[str], values) -> dict[str, str]:
    return {header: clean_cell(value) for header, value in zip(headers, values) if header}


def normalized_status(value: str) -> str:
    text = value.strip().lower()
    if text in {"maintenance", "维护", "维护中", "检修", "检修中"}:
        return Instrument.Status.MAINTENANCE
    if text in {"disabled", "停用", "暂停使用", "报废"}:
        return Instrument.Status.DISABLED
    return Instrument.Status.NORMAL


def parse_assignment_rows(workbook) -> list[InstrumentImportRow]:
    sheet = sheet_by_keywords(workbook, ["分工", "台账"]) or workbook.worksheets[0]
    headers = [clean_cell(value) for value in next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))]
    name_counts: dict[str, int] = defaultdict(int)
    rows: list[InstrumentImportRow] = []

    for values in sheet.iter_rows(min_row=2, values_only=True):
        data = row_dict(headers, values)
        name = data.get("设备名称") or data.get("仪器名称") or data.get("名称") or ""
        if not name:
            continue

        name_counts[name] += 1
        index_text = data.get("序号") or str(len(rows) + 1)
        try:
            index = int(float(index_text))
        except ValueError:
            index = len(rows) + 1

        rows.append(
            InstrumentImportRow(
                index=index,
                name=name,
                model=data.get("型号", "") or data.get("设备型号", "") or data.get("仪器型号", ""),
                owner=data.get("设备归属", ""),
                is_asset=data.get("是否国资", ""),
                location_detail=data.get("详细位置", "") or data.get("存放位置", "") or data.get("房间", ""),
                remark=data.get("厂家/年限/备注", "") or data.get("备注", ""),
                manager_name=data.get("管理员", "") or data.get("负责人", ""),
                notes=data.get("使用说明", "") or data.get("说明", ""),
                status=normalized_status(data.get("状态", "")),
                occurrence=name_counts[name],
            )
        )
    return rows


def image_extension(image) -> str:
    fmt = (getattr(image, "format", "") or "").lower()
    if fmt in {"jpeg", "jpg"}:
        return "jpg"
    if fmt in {"png", "gif", "webp"}:
        return fmt
    suffix = Path(getattr(image, "path", "") or "").suffix.lower().lstrip(".")
    return suffix or "png"


def parse_images(workbook) -> dict[tuple[str, int], tuple[str, bytes]]:
    sheet = sheet_by_keywords(workbook, ["图片", "照片"])
    if sheet is None:
        assignment_sheet = sheet_by_keywords(workbook, ["分工", "台账"]) or workbook.worksheets[0]
        sheet = assignment_sheet if getattr(assignment_sheet, "_images", []) else None
    if sheet is None:
        return {}

    headers = [clean_cell(value) for value in next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))]
    names_by_row: dict[int, str] = {}
    for row_number, values in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        data = row_dict(headers, values)
        name = data.get("设备名称") or data.get("仪器名称") or data.get("名称") or ""
        if name:
            names_by_row[row_number] = name

    occurrence_by_name: dict[str, int] = defaultdict(int)
    images: dict[tuple[str, int], tuple[str, bytes]] = {}
    for image in getattr(sheet, "_images", []):
        marker = getattr(image.anchor, "_from", None)
        if marker is None:
            continue
        row_number = marker.row + 1
        name = names_by_row.get(row_number)
        if not name:
            continue
        occurrence_by_name[name] += 1
        ext = image_extension(image)
        images[(name, occurrence_by_name[name])] = (f"{slugify(name, allow_unicode=True) or 'instrument'}-{occurrence_by_name[name]}.{ext}", image._data())
    return images


def instrument_notes(row: InstrumentImportRow) -> str:
    if row.notes:
        return row.notes
    lines = []
    if row.owner:
        lines.append(f"设备归属：{row.owner}")
    if row.is_asset:
        lines.append(f"是否国资：{row.is_asset}")
    if row.manager_name:
        lines.append(f"管理员：{row.manager_name}")
    if row.remark:
        lines.append(f"厂家/年限/备注：{row.remark}")
    return "\n".join(lines)


def find_manager(name: str):
    if not name:
        return None
    User = get_user_model()
    return (
        User.objects.select_related("profile")
        .filter(profile__real_name=name)
        .first()
        or User.objects.filter(first_name=name).first()
        or User.objects.filter(username=name).first()
    )


@transaction.atomic
def import_instruments_from_excel(file_obj: BinaryIO, *, uploaded_by=None) -> dict[str, int]:
    from openpyxl import load_workbook

    workbook = load_workbook(file_obj, data_only=True)
    rows = parse_assignment_rows(workbook)
    images = parse_images(workbook)
    category, _ = InstrumentCategory.objects.get_or_create(
        slug=DEFAULT_CATEGORY_SLUG,
        defaults={"name": DEFAULT_CATEGORY_NAME, "description": "课题组实验室仪器与设备。", "sort_order": 10},
    )

    created = 0
    updated = 0
    image_count = 0
    existing_by_name: dict[str, list[Instrument]] = defaultdict(list)
    for instrument in Instrument.objects.all().order_by("sort_order", "id"):
        existing_by_name[instrument.name].append(instrument)

    for row in rows:
        manager = find_manager(row.manager_name)
        defaults = {
            "name": row.name,
            "model": row.model,
            "category": category,
            "location_detail": row.location_detail,
            "manager": manager,
            "status": row.status,
            "notes": instrument_notes(row),
            "sort_order": row.index,
        }
        matches = existing_by_name[row.name]
        if row.occurrence <= len(matches):
            instrument = matches[row.occurrence - 1]
            for field, value in defaults.items():
                setattr(instrument, field, value)
            instrument.save()
            was_created = False
        else:
            instrument = Instrument.objects.create(**defaults)
            matches.append(instrument)
            was_created = True
        created += int(was_created)
        updated += int(not was_created)

        image_data = images.get((row.name, row.occurrence))
        if image_data:
            filename, content = image_data
            instrument.image.save(filename, ContentFile(content), save=True)
            image_count += 1

    return {"created": created, "updated": updated, "images": image_count, "total": len(rows)}
