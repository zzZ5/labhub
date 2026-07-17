from io import BytesIO

import pytest
from django.contrib.auth import get_user_model
from django.core.files.base import ContentFile
from django.urls import reverse
from openpyxl import Workbook
from openpyxl.drawing.image import Image as WorksheetImage
from PIL import Image as PillowImage
from apps.accounts.models import Role, RoleCode, UserRole

from .models import Instrument, InstrumentCategory

User = get_user_model()


@pytest.fixture
def member(db):
    user = User.objects.create_user(username="member", password="pass12345")
    user.profile.is_approved = True
    user.profile.save()
    role, _ = Role.objects.get_or_create(code=RoleCode.MEMBER, defaults={"name": "课题组成员"})
    UserRole.objects.get_or_create(user=user, role=role)
    return user


@pytest.fixture
def manager(db):
    user = User.objects.create_user(username="manager", password="pass12345")
    user.profile.is_approved = True
    user.profile.save()
    role, _ = Role.objects.get_or_create(code=RoleCode.INSTRUMENT_MANAGER, defaults={"name": "仪器管理员"})
    UserRole.objects.get_or_create(user=user, role=role)
    return user


@pytest.fixture
def instrument(db, manager):
    category = InstrumentCategory.objects.create(name="分析仪器", slug="analysis")
    return Instrument.objects.create(
        name="总有机碳分析仪",
        category=category,
        manager=manager,
        location_detail="B112",
        notes="使用前确认载气压力。",
    )


@pytest.mark.django_db
def test_instrument_list_requires_login(client):
    response = client.get(reverse("instrument-list"))

    assert response.status_code in {401, 403}


@pytest.mark.django_db
def test_member_can_view_instrument_details(client, member, instrument):
    client.login(username="member", password="pass12345")

    response = client.get(reverse("instrument-list"))

    assert response.status_code == 200
    assert response.json()[0]["name"] == "总有机碳分析仪"
    assert response.json()[0]["location_detail"] == "B112"
    assert response.json()[0]["notes"] == "使用前确认载气压力。"
    assert "need_training" not in response.json()[0]
    assert "serial_number" not in response.json()[0]


@pytest.mark.django_db
def test_member_cannot_edit_instrument(client, member, instrument):
    client.login(username="member", password="pass12345")

    response = client.patch(
        reverse("instrument-detail", args=[instrument.id]),
        {"name": "被越权修改的名称"},
        content_type="application/json",
    )

    assert response.status_code == 403
    instrument.refresh_from_db()
    assert instrument.name == "总有机碳分析仪"


@pytest.mark.django_db
def test_instrument_manager_can_update_core_fields(client, manager, instrument):
    client.login(username="manager", password="pass12345")

    response = client.patch(
        reverse("instrument-detail", args=[instrument.id]),
        {
            "location_detail": "环境分析实验室 B113",
            "status": Instrument.Status.MAINTENANCE,
            "notes": "维护期间暂停使用。",
        },
        content_type="application/json",
    )

    assert response.status_code == 200
    instrument.refresh_from_db()
    assert instrument.location_detail == "环境分析实验室 B113"
    assert instrument.status == Instrument.Status.MAINTENANCE


@pytest.mark.django_db
def test_removed_instrument_workflow_endpoints_return_not_found(client, manager):
    client.login(username="manager", password="pass12345")

    for path in ("training-records/", "maintenance-records/", "fault-reports/"):
        response = client.get(f"/api/instruments/{path}")
        assert response.status_code == 404


@pytest.mark.django_db
def test_instrument_import_updates_matching_name_without_duplicate(client, manager, instrument):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "仪器分工表"
    sheet.append(["序号", "仪器名称", "型号", "状态", "详细位置", "使用说明"])
    sheet.append([1, "总有机碳分析仪", "TOC-L", "维护中", "环境分析实验室 B113", "校准后使用。"])
    buffer = BytesIO()
    workbook.save(buffer)
    client.force_login(manager)

    response = client.post(
        reverse("instrument-import-excel"),
        {"file": ContentFile(buffer.getvalue(), name="instruments.xlsx")},
    )

    assert response.status_code == 200
    assert response.json()["created"] == 0
    assert response.json()["updated"] == 1
    assert Instrument.objects.filter(name="总有机碳分析仪").count() == 1
    instrument.refresh_from_db()
    assert instrument.model == "TOC-L"
    assert instrument.status == Instrument.Status.MAINTENANCE
    assert instrument.location_detail == "环境分析实验室 B113"


@pytest.mark.django_db
def test_instrument_import_keeps_embedded_equipment_image(client, manager):
    workbook = Workbook()
    assignment = workbook.active
    assignment.title = "仪器分工表"
    assignment.append(["序号", "仪器名称", "型号", "状态", "详细位置", "使用说明"])
    assignment.append([1, "堆肥呼吸仪", "CR-01", "正常", "环境楼 201", "使用后清洁气路。"])
    images = workbook.create_sheet("仪器图片")
    images.append(["仪器名称", "图片"])
    images.append(["堆肥呼吸仪", ""])
    image_buffer = BytesIO()
    PillowImage.new("RGB", (12, 9), color=(0, 135, 60)).save(image_buffer, format="PNG")
    image_buffer.seek(0)
    images.add_image(WorksheetImage(PillowImage.open(image_buffer)), "B2")
    workbook_buffer = BytesIO()
    workbook.save(workbook_buffer)
    client.force_login(manager)

    response = client.post(
        reverse("instrument-import-excel"),
        {"file": ContentFile(workbook_buffer.getvalue(), name="instruments-with-image.xlsx")},
    )

    assert response.status_code == 200
    assert response.json()["created"] == 1
    assert response.json()["images"] == 1
    instrument = Instrument.objects.get(name="堆肥呼吸仪")
    assert instrument.image
    assert instrument.image.size > 0
