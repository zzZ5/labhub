import mimetypes
from pathlib import Path
from zipfile import BadZipFile, ZipFile

from django.core.files.base import ContentFile
from django.db import transaction
from django.utils import timezone
from openpyxl import load_workbook

from apps.students.preview import refresh_file_preview_pdf
from apps.system.uploads import validate_document_upload, validate_spreadsheet_upload

from .models import Document, DocumentCategory, DocumentStatus
from .services import can_edit_document


HEADER_ALIASES = {
    "title": ["资料标题", "标题", "名称"],
    "category": ["资料分类", "分类", "分类名称"],
    "description": ["资料说明", "说明", "简介", "备注"],
    "filename": ["文件名", "附件文件名", "zip内文件名", "ZIP内文件名"],
    "allow_download": ["允许下载", "是否允许下载", "下载"],
}

def cell_text(value):
    if value is None:
        return ""
    return str(value).strip()


def normalize_key(value):
    return cell_text(value).replace(" ", "").replace("\n", "").lower()


def truthy(value, default=True):
    text = normalize_key(value)
    if not text:
        return default
    if text in {"是", "允许", "true", "yes", "y", "1"}:
        return True
    if text in {"否", "不允许", "false", "no", "n", "0"}:
        return False
    return default


def header_map(row):
    normalized = {normalize_key(value): index for index, value in enumerate(row)}
    result = {}
    for field, aliases in HEADER_ALIASES.items():
        for alias in aliases:
            key = normalize_key(alias)
            if key in normalized:
                result[field] = normalized[key]
                break
    return result


def row_value(row, mapping, field):
    index = mapping.get(field)
    if index is None or index >= len(row):
        return ""
    return cell_text(row[index])


def get_data_sheet(workbook):
    if "导入数据" in workbook.sheetnames:
        return workbook["导入数据"]
    if len(workbook.worksheets) >= 2:
        return workbook.worksheets[1]
    return workbook.worksheets[0]


def category_lookup():
    categories = DocumentCategory.objects.all()
    by_name = {item.name.strip(): item for item in categories}
    by_slug = {item.slug.strip(): item for item in categories}
    return by_name, by_slug


def resolve_category(value, by_name, by_slug):
    text = cell_text(value)
    if not text:
        return None
    return by_slug.get(text) or by_name.get(text)


def archive_files(archive_file):
    if not archive_file:
        return {}
    validate_document_upload(archive_file)
    files = {}
    try:
        with ZipFile(archive_file) as archive:
            for info in archive.infolist():
                if info.is_dir():
                    continue
                name = Path(info.filename.replace("\\", "/")).name
                if not name or name.startswith("."):
                    continue
                files[name] = archive.read(info)
    except BadZipFile as exc:
        raise ValueError("文件包不是有效的 zip 文件。") from exc
    finally:
        archive_file.seek(0)
    return files


def create_or_update_document(row, mapping, *, categories_by_name, categories_by_slug, files, user):
    title = row_value(row, mapping, "title")
    if not title:
        return "skipped", "资料标题为空，已跳过。"

    category_text = row_value(row, mapping, "category")
    category = resolve_category(category_text, categories_by_name, categories_by_slug)
    if category_text and not category:
        return "error", f"资料分类不存在：{category_text}"

    filename = row_value(row, mapping, "filename")
    file_bytes = None
    if filename:
        filename = Path(filename.replace("\\", "/")).name
        file_bytes = files.get(filename)
        if file_bytes is None:
            return "error", f"zip 文件包中找不到：{filename}"

    defaults = {
        "category": category,
        "description": row_value(row, mapping, "description"),
        "allow_download": truthy(row_value(row, mapping, "allow_download"), default=True),
        "status": DocumentStatus.ACTIVE,
    }
    document = Document.objects.filter(title=title).first()
    created = document is None
    if document and not can_edit_document(user, document):
        return "error", "已有同名资料，但当前账号无权更新。"

    if created:
        document = Document.objects.create(title=title, owner=user, uploaded_by=user, **defaults)

    if not created:
        for field, value in defaults.items():
            setattr(document, field, value)
        document.save()

    if file_bytes is not None:
        old_file = document.file.name if document.file else ""
        old_preview = document.preview_pdf.name if document.preview_pdf else ""
        content_type = mimetypes.guess_type(filename)[0] or ""
        file_content = ContentFile(file_bytes, name=filename)
        validate_document_upload(file_content)
        document.file.save(filename, file_content, save=False)
        document.preview_pdf = ""
        document.preview_status = "none"
        document.preview_error = ""
        document.original_filename = filename
        document.uploaded_by = user
        document.uploaded_at = timezone.now()
        document.file_size = len(file_bytes)
        document.file_type = content_type
        document.save()
        if old_file and old_file != document.file.name:
            document.file.storage.delete(old_file)
        if old_preview:
            document.preview_pdf.storage.delete(old_preview)
        refresh_file_preview_pdf(document)

    return ("created" if created else "updated"), ""


@transaction.atomic
def import_documents_excel(excel_file, *, archive_file=None, user=None):
    if not excel_file.name.lower().endswith(".xlsx"):
        raise ValueError("请上传 .xlsx 格式的资料清单。")
    validate_spreadsheet_upload(excel_file)

    workbook = load_workbook(excel_file, data_only=True)
    sheet = get_data_sheet(workbook)
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return {"created": 0, "updated": 0, "skipped": 0, "errors": ["导入表为空。"]}

    mapping = header_map(rows[0])
    if "title" not in mapping:
        return {"created": 0, "updated": 0, "skipped": 0, "errors": ["导入表缺少“资料标题”列。"]}

    files = archive_files(archive_file)
    categories_by_name, categories_by_slug = category_lookup()
    result = {"created": 0, "updated": 0, "skipped": 0, "errors": []}

    for row_number, row in enumerate(rows[1:], start=2):
        if not any(cell_text(value) for value in row):
            continue
        action, message = create_or_update_document(
            row,
            mapping,
            categories_by_name=categories_by_name,
            categories_by_slug=categories_by_slug,
            files=files,
            user=user,
        )
        if action in {"created", "updated", "skipped"}:
            result[action] += 1
        if message:
            result["errors"].append(f"第 {row_number} 行：{message}")

    return result
