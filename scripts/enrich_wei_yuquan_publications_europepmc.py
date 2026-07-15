from __future__ import annotations

import json
import re
import time
import urllib.parse
import urllib.request
from difflib import SequenceMatcher
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
IMPORT_DIR = ROOT / "outputs" / "portal-import" / "wei-yuquan-2026-7"
PUBLICATION_FILE = IMPORT_DIR / "05-publications-wei-yuquan.xlsx"
CACHE_FILE = IMPORT_DIR / "publication_europepmc_cache.json"
REPORT_FILE = IMPORT_DIR / "publication_europepmc_report.md"


def normalize_title(value: str) -> str:
    text = value.lower()
    text = re.sub(r"[^a-z0-9\u4e00-\u9fff]+", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def similarity(a: str, b: str) -> float:
    left = normalize_title(a)
    right = normalize_title(b)
    if not left or not right:
        return 0.0
    if left == right:
        return 1.0
    if left in right or right in left:
        return 0.94
    return SequenceMatcher(None, left, right).ratio()


def fetch_json(url: str) -> dict:
    request = urllib.request.Request(
        url,
        headers={
            "User-Agent": "LabHub metadata enrichment (mailto:weiyq2019@cau.edu.cn)",
            "Accept": "application/json",
        },
    )
    with urllib.request.urlopen(request, timeout=25) as response:
        return json.loads(response.read().decode("utf-8"))


def clean_abstract(text: str | None) -> str:
    if not text:
        return ""
    text = re.sub(r"<[^>]+>", " ", text)
    text = re.sub(r"\s+", " ", text).strip()
    return text


def query_europepmc(title: str, year: str, doi: str) -> dict:
    query = f"DOI:{doi}" if doi else f'TITLE:"{title}"'
    params = {
        "query": query,
        "format": "json",
        "pageSize": "5",
        "resultType": "core",
    }
    url = "https://www.ebi.ac.uk/europepmc/webservices/rest/search?" + urllib.parse.urlencode(params)
    data = fetch_json(url)
    best = None
    best_score = 0.0
    for item in data.get("resultList", {}).get("result", []):
        item_title = item.get("title") or ""
        score = 1.0 if doi and (item.get("doi") or "").lower() == doi.lower() else similarity(title, item_title)
        item_year = str(item.get("pubYear") or "")
        if year and item_year and item_year != year:
            score -= 0.05
        if score > best_score:
            best = item
            best_score = score
    if not best or best_score < 0.72:
        return {"matched": False, "score": round(best_score, 3)}
    return {
        "matched": True,
        "score": round(best_score, 3),
        "source": "Europe PMC",
        "title": best.get("title") or "",
        "doi": best.get("doi") or "",
        "journal": best.get("journalTitle") or "",
        "year": str(best.get("pubYear") or ""),
        "volume": best.get("journalVolume") or "",
        "issue": best.get("issue") or "",
        "pages": best.get("pageInfo") or "",
        "abstract": clean_abstract(best.get("abstractText")),
    }


def main() -> None:
    cache = json.loads(CACHE_FILE.read_text(encoding="utf-8")) if CACHE_FILE.exists() else {}
    workbook = load_workbook(PUBLICATION_FILE)
    sheet = workbook["Data"]
    headers = [cell.value for cell in sheet[1]]
    columns = {header: index + 1 for index, header in enumerate(headers)}

    attempted = matched = added_abstracts = added_dois = 0
    errors: list[str] = []
    for row_index in range(2, sheet.max_row + 1):
        title = str(sheet.cell(row_index, columns["title"]).value or "").strip()
        year = str(sheet.cell(row_index, columns["year"]).value or "").strip()
        doi = str(sheet.cell(row_index, columns["doi"]).value or "").strip()
        abstract = str(sheet.cell(row_index, columns["abstract"]).value or "").strip()
        if not title or abstract:
            continue
        if not doi and not re.search(r"[A-Za-z]", title):
            continue

        key = f"{doi or title}::{year}"
        if key not in cache:
            attempted += 1
            try:
                cache[key] = query_europepmc(title, year, doi)
            except Exception as exc:
                cache[key] = {"matched": False, "error": str(exc)}
                errors.append(f"{title}: {exc}")
            CACHE_FILE.write_text(json.dumps(cache, ensure_ascii=False, indent=2), encoding="utf-8")
            time.sleep(0.25)

        item = cache[key]
        if item.get("matched"):
            matched += 1
        if item.get("doi") and not doi:
            sheet.cell(row_index, columns["doi"]).value = item["doi"]
            added_dois += 1
        abstract_text = str(item.get("abstract") or "").strip()
        if len(abstract_text) >= 40:
            sheet.cell(row_index, columns["abstract"]).value = abstract_text
            added_abstracts += 1
        for field in ["journal", "volume", "issue", "pages"]:
            value = str(item.get(field) or "").strip()
            column = columns.get(field)
            current = str(sheet.cell(row_index, column).value or "").strip() if column else ""
            if column and value and not current:
                sheet.cell(row_index, column).value = value

    workbook.save(PUBLICATION_FILE)
    REPORT_FILE.write_text(
        "\n".join(
            [
                "# Europe PMC 补全报告",
                "",
                f"- 本轮新查询：{attempted}",
                f"- 命中记录：{matched}",
                f"- 新增 DOI：{added_dois}",
                f"- 新增摘要：{added_abstracts}",
                f"- 错误数：{len(errors)}",
                "",
                "## 错误样例",
                *(f"- {line}" for line in errors[:30]),
            ]
        ),
        encoding="utf-8",
    )
    print({"attempted": attempted, "matched": matched, "added_dois": added_dois, "added_abstracts": added_abstracts, "errors": len(errors)})


if __name__ == "__main__":
    main()
