from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


OUTPUT = Path(__file__).resolve().parents[1] / "frontend" / "public" / "templates" / "documents-import-template.xlsx"


def style_sheet(sheet):
    sheet.freeze_panes = "A2"
    header_fill = PatternFill("solid", fgColor="00873C")
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def main():
    workbook = Workbook()
    guide = workbook.active
    guide.title = "导入说明"
    guide["A1"] = "内部资料批量导入说明"
    guide["A1"].font = Font(size=16, bold=True, color="1F3D2B")
    guide["A3"] = "填写方式"
    guide["B3"] = "请在第二张表“导入数据”填写资料。需要带文件时，将 PDF、Word、PPT、Excel、图片等文件打包为 zip，并在“文件名”列填写 zip 内对应文件名。"
    guide["A4"] = "分类"
    guide["B4"] = "建议填写：组内制度与通知、实验室安全、实验方法、项目与经费材料、论文写作与投稿、其他参考资料、组会与学术交流、行政表格与模板。也可以填写分类 slug。"
    guide["A5"] = "允许下载"
    guide["B5"] = "可填写：是/否。留空默认允许下载。"
    guide["A6"] = "重复标题"
    guide["B6"] = "标题重复时会更新原资料；普通成员只能更新自己上传的资料。"
    guide.column_dimensions["A"].width = 18
    guide.column_dimensions["B"].width = 96
    for row in range(3, 7):
        guide[f"A{row}"].font = Font(bold=True, color="1F3D2B")
        guide[f"A{row}"].fill = PatternFill("solid", fgColor="EAF5EE")
        guide[f"A{row}"].alignment = Alignment(vertical="top")
        guide[f"B{row}"].alignment = Alignment(vertical="top", wrap_text=True)

    data = workbook.create_sheet("导入数据")
    headers = ["资料标题", "资料分类", "资料说明", "文件名", "允许下载"]
    examples = [
        ["堆肥样品前处理流程", "实验方法", "样品采集、保存、前处理和记录要求。", "堆肥样品前处理流程.pdf", "是"],
        ["组会汇报模板", "组会与学术交流", "组会汇报 PPT 模板。", "组会汇报模板.pptx", "是"],
        ["项目报销材料清单", "项目与经费材料", "项目经费材料准备参考。", "项目报销材料清单.docx", "否"],
        ["有机废弃物资源化研究学位论文", "其他参考资料", "堆肥与资源化方向参考资料。", "有机废弃物资源化研究学位论文.pdf", "是"],
    ]
    data.append(headers)
    for row in examples:
        data.append(row)
    style_sheet(data)
    widths = [28, 20, 42, 34, 14]
    for index, width in enumerate(widths, start=1):
        data.column_dimensions[get_column_letter(index)].width = width
    for row in range(2, 80):
        data.row_dimensions[row].height = 42

    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(OUTPUT)
    print(OUTPUT)


if __name__ == "__main__":
    main()
