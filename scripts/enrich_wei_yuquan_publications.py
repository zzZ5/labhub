from __future__ import annotations

import json
import re
import time
import urllib.parse
import urllib.request
from difflib import SequenceMatcher
from html import unescape
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
IMPORT_DIR = ROOT / "outputs" / "portal-import" / "wei-yuquan-2026-7"
PUBLICATION_FILE = IMPORT_DIR / "05-publications-wei-yuquan.xlsx"
CACHE_FILE = IMPORT_DIR / "publication_enrichment_cache.json"
REPORT_FILE = IMPORT_DIR / "publication_enrichment_report.md"


def normalize_title(value: str) -> str:
    text = value.lower()
    text = re.sub(r"<[^>]+>", " ", text)
    text = re.sub(r"[^a-z0-9\u4e00-\u9fff]+", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def similarity(a: str, b: str) -> float:
    left = normalize_title(a)
    right = normalize_title(b)
    if not left or not right:
        return 0.0
    if left == right:
        return 1.0
    if left in right or right in left:
        return 0.94
    return SequenceMatcher(None, left, right).ratio()


def abstract_from_inverted_index(index: dict | None) -> str:
    if not index:
        return ""
    positions: list[tuple[int, str]] = []
    for word, indexes in index.items():
        for position in indexes:
            positions.append((int(position), word))
    return " ".join(word for _, word in sorted(positions)).strip()


def strip_abstract(value: str | None) -> str:
    if not value:
        return ""
    text = unescape(value)
    text = re.sub(r"<[^>]+>", " ", text)
    text = re.sub(r"\b(Abstract|ABSTRACT)\b[:：]?", "", text).strip()
    text = re.sub(r"\s+", " ", text)
    return text.strip()


def fetch_json(url: str) -> dict:
    request = urllib.request.Request(
        url,
        headers={
            "User-Agent": "LabHub metadata enrichment (mailto:weiyq2019@cau.edu.cn)",
            "Accept": "application/json",
        },
    )
    with urllib.request.urlopen(request, timeout=30) as response:
        return json.loads(response.read().decode("utf-8"))


def openalex_lookup(title: str, year: str) -> dict:
    params = {
        "search": title,
        "per-page": "5",
        "select": "id,doi,title,display_name,publication_year,abstract_inverted_index,primary_location,biblio",
    }
    url = "https://api.openalex.org/works?" + urllib.parse.urlencode(params)
    data = fetch_json(url)
    best: dict | None = None
    best_score = 0.0
    for item in data.get("results", []):
        candidate_title = item.get("title") or item.get("display_name") or ""
        score = similarity(title, candidate_title)
        item_year = str(item.get("publication_year") or "")
        if year and item_year and item_year != year:
            score -= 0.08
        if score > best_score:
            best = item
            best_score = score
    if not best or best_score < 0.72:
        return {"matched": False, "score": round(best_score, 3)}

    location = best.get("primary_location") or {}
    source = location.get("source") or {}
    biblio = best.get("biblio") or {}
    doi = (best.get("doi") or "").replace("https://doi.org/", "")
    return {
        "matched": True,
        "score": round(best_score, 3),
        "source": "OpenAlex",
        "openalex_id": best.get("id") or "",
        "title": best.get("title") or best.get("display_name") or "",
        "doi": doi,
        "journal": source.get("display_name") or "",
        "year": str(best.get("publication_year") or ""),
        "volume": biblio.get("volume") or "",
        "issue": biblio.get("issue") or "",
        "pages": biblio.get("first_page") or "",
        "abstract": abstract_from_inverted_index(best.get("abstract_inverted_index")),
    }


def crossref_lookup(title: str, year: str) -> dict:
    params = {
        "query.title": title,
        "rows": "5",
        "select": "DOI,title,container-title,published-print,published-online,issued,volume,issue,page,abstract",
    }
    url = "https://api.crossref.org/works?" + urllib.parse.urlencode(params)
    data = fetch_json(url)
    best: dict | None = None
    best_score = 0.0
    for item in data.get("message", {}).get("items", []):
        titles = item.get("title") or []
        candidate_title = titles[0] if titles else ""
        score = similarity(title, candidate_title)
        item_year = crossref_year(item)
        if year and item_year and item_year != year:
            score -= 0.08
        if score > best_score:
            best = item
            best_score = score
    if not best or best_score < 0.72:
        return {"matched": False, "score": round(best_score, 3)}

    journals = best.get("container-title") or []
    return {
        "matched": True,
        "score": round(best_score, 3),
        "source": "Crossref",
        "doi": best.get("DOI") or "",
        "journal": journals[0] if journals else "",
        "year": crossref_year(best),
        "volume": best.get("volume") or "",
        "issue": best.get("issue") or "",
        "pages": best.get("page") or "",
        "abstract": strip_abstract(best.get("abstract")),
    }


def crossref_year(item: dict) -> str:
    for key in ["published-print", "published-online", "issued"]:
        parts = (((item.get(key) or {}).get("date-parts") or [[]])[0] or [])
        if parts:
            return str(parts[0])
    return ""


def enrich_one(title: str, year: str) -> dict:
    try:
        result = openalex_lookup(title, year)
        if result.get("matched") and (result.get("abstract") or result.get("doi")):
            return result
    except Exception as exc:
        result = {"matched": False, "error": f"OpenAlex: {exc}"}

    time.sleep(0.25)
    try:
        fallback = crossref_lookup(title, year)
        if fallback.get("matched"):
            if not fallback.get("abstract") and result.get("matched"):
                fallback["abstract"] = result.get("abstract", "")
            return fallback
        return result if result.get("matched") else fallback
    except Exception as exc:
        if result.get("matched"):
            result["crossref_error"] = str(exc)
            return result
        return {"matched": False, "error": f"Crossref: {exc}"}


def main() -> None:
    cache = json.loads(CACHE_FILE.read_text(encoding="utf-8")) if CACHE_FILE.exists() else {}
    workbook = load_workbook(PUBLICATION_FILE)
    sheet = workbook["Data"]
    headers = [cell.value for cell in sheet[1]]
    columns = {header: index + 1 for index, header in enumerate(headers)}

    updated_abstracts = 0
    updated_dois = 0
    matched = 0
    unmatched: list[str] = []

    for row_index in range(2, sheet.max_row + 1):
        title = str(sheet.cell(row_index, columns["title"]).value or "").strip()
        year = str(sheet.cell(row_index, columns["year"]).value or "").strip()
        if not title:
            continue

        key = f"{title}::{year}"
        if key not in cache:
            cache[key] = enrich_one(title, year)
            CACHE_FILE.write_text(json.dumps(cache, ensure_ascii=False, indent=2), encoding="utf-8")
            time.sleep(0.35)

        item = cache[key]
        if item.get("matched"):
            matched += 1
        else:
            unmatched.append(title)

        doi = str(item.get("doi") or "").strip()
        if doi and not str(sheet.cell(row_index, columns["doi"]).value or "").strip():
            sheet.cell(row_index, columns["doi"]).value = doi
            updated_dois += 1

        abstract = str(item.get("abstract") or "").strip()
        if abstract and len(abstract) >= 40:
            sheet.cell(row_index, columns["abstract"]).value = abstract
            updated_abstracts += 1

        for field in ["journal", "volume", "issue", "pages"]:
            value = str(item.get(field) or "").strip()
            column = columns.get(field)
            current = str(sheet.cell(row_index, column).value or "").strip() if column else ""
            if column and value and not current:
                sheet.cell(row_index, column).value = value

    workbook.save(PUBLICATION_FILE)
    REPORT_FILE.write_text(
        "\n".join(
            [
                "# 论文元数据补全报告",
                "",
                f"- 论文总数：{sheet.max_row - 1}",
                f"- 匹配到外部记录：{matched}",
                f"- 新增 DOI：{updated_dois}",
                f"- 写入真实摘要：{updated_abstracts}",
                f"- 未匹配：{len(unmatched)}",
                "",
                "## 未匹配题名",
                *(f"- {title}" for title in unmatched[:80]),
            ]
        ),
        encoding="utf-8",
    )
    print({"total": sheet.max_row - 1, "matched": matched, "updated_dois": updated_dois, "updated_abstracts": updated_abstracts, "unmatched": len(unmatched)})


if __name__ == "__main__":
    main()
