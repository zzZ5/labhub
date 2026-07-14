from __future__ import annotations

from collections import defaultdict
from io import BytesIO
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.datavalidation import DataValidation


ROOT = Path(__file__).resolve().parents[1]
SOURCE = ROOT / "实验室设备管理人员0710.xlsx"
OUTPUT_DIR = ROOT / "outputs" / "instruments"
PUBLIC_TEMPLATE_DIR = ROOT / "frontend" / "public" / "templates"

HEADERS = ["仪器名称", "型号", "状态", "详细位置", "设备图片", "使用说明"]
STATUS_OPTIONS = "正常,维护中,停用"
HEADER_FILL = PatternFill("solid", fgColor="00873C")
HEADER_FONT = Font(color="FFFFFF", bold=True)
THIN = Side(style="thin", color="D9E2DC")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

HELP_ROWS = [
    ["仪器平台批量导入说明"],
    ["1. “仪器台账”只保留 6 列：仪器名称、型号、状态、详细位置、设备图片、使用说明。"],
    ["2. 仪器名称必填；状态可填：正常、维护中、停用；型号、详细位置、使用说明可后续补充。"],
    ["3. 设备图片直接插入到对应行的“设备图片”列即可，系统会按图片所在行匹配仪器。"],
    ["4. 同名仪器会按出现顺序区分；重复导入时，系统按行序更新已有记录，不会反复新增。"],
]


def clean_cell(value) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    if text.endswith(".0") and text[:-2].isdigit():
        return text[:-2]
    return text


def style_header(cell):
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = BORDER


def style_sheet(sheet):
    sheet.sheet_view.showGridLines = False
    sheet.freeze_panes = "A2"
    for cell in sheet[1]:
        style_header(cell)
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.border = BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    sheet.row_dimensions[1].height = 26


def add_status_validation(sheet):
    validation = DataValidation(type="list", formula1=f'"{STATUS_OPTIONS}"', allow_blank=False)
    sheet.add_data_validation(validation)
    validation.add("C2:C300")


def build_help_sheet(workbook):
    sheet = workbook.create_sheet("导入说明", 0)
    for row in HELP_ROWS:
        sheet.append(row)
    style_sheet(sheet)
    sheet["A1"].font = Font(color="FFFFFF", bold=True, size=14)
    sheet.column_dimensions["A"].width = 100
    for row_index in range(1, len(HELP_ROWS) + 1):
        sheet.row_dimensions[row_index].height = 30 if row_index > 1 else 28


def source_rows(workbook):
    sheet = workbook["分工表"]
    headers = [clean_cell(cell.value) for cell in sheet[1]]
    rows = []
    for values in sheet.iter_rows(min_row=2, values_only=True):
        data = {header: clean_cell(value) for header, value in zip(headers, values) if header}
        name = data.get("设备名称", "")
        if not name:
            continue
        rows.append(
            {
                "name": name,
                "model": "",
                "status": "正常",
                "location": data.get("存放位置", ""),
                "notes": data.get("厂家/年限/备注", ""),
            }
        )
    return rows


def source_images(workbook):
    sheet = workbook["仪器图片"]
    headers = [clean_cell(cell.value) for cell in sheet[1]]
    names_by_row = {}
    for row_number, values in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        data = {header: clean_cell(value) for header, value in zip(headers, values) if header}
        if data.get("设备名称"):
            names_by_row[row_number] = data["设备名称"]

    occurrence_by_name = defaultdict(int)
    images = {}
    for image in getattr(sheet, "_images", []):
        marker = getattr(image.anchor, "_from", None)
        if marker is None:
            continue
        name = names_by_row.get(marker.row + 1)
        if not name:
            continue
        occurrence_by_name[name] += 1
        images[(name, occurrence_by_name[name])] = image._data()
    return images


def add_scaled_image(sheet, cell, image_bytes: bytes):
    image = ExcelImage(BytesIO(image_bytes))
    max_width = 150
    max_height = 115
    scale = min(max_width / image.width, max_height / image.height, 1)
    image.width = int(image.width * scale)
    image.height = int(image.height * scale)
    sheet.add_image(image, cell)


def save_workbook(workbook, output_path: Path) -> Path:
    try:
        workbook.save(output_path)
        return output_path
    except PermissionError:
        fallback = output_path.with_name(f"{output_path.stem}-v2{output_path.suffix}")
        workbook.save(fallback)
        return fallback


def create_simplified_workbook():
    source = load_workbook(SOURCE, data_only=True)
    rows = source_rows(source)
    images = source_images(source)
    occurrence_by_name = defaultdict(int)

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "仪器台账"
    sheet.append(HEADERS)

    for row in rows:
        sheet.append([row["name"], row["model"], row["status"], row["location"], "", row["notes"]])

    style_sheet(sheet)
    add_status_validation(sheet)
    widths = {"A": 24, "B": 16, "C": 12, "D": 22, "E": 24, "F": 42}
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width
    for row_index in range(2, sheet.max_row + 1):
        sheet.row_dimensions[row_index].height = 92
        name = clean_cell(sheet.cell(row_index, 1).value)
        occurrence_by_name[name] += 1
        image_bytes = images.get((name, occurrence_by_name[name]))
        if image_bytes:
            add_scaled_image(sheet, f"E{row_index}", image_bytes)

    build_help_sheet(workbook)

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    output_path = OUTPUT_DIR / "实验室仪器平台导入简化版.xlsx"
    return save_workbook(workbook, output_path)


def create_public_template():
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "仪器台账"
    sheet.append(HEADERS)
    sheet.append(["示例：电热恒温培养箱", "", "正常", "沃土实验室一楼", "请把图片插入到这一格附近", "使用说明、注意事项或厂家购置备注。"])
    style_sheet(sheet)
    add_status_validation(sheet)
    for column, width in {"A": 24, "B": 16, "C": 12, "D": 22, "E": 28, "F": 42}.items():
        sheet.column_dimensions[column].width = width
    sheet.row_dimensions[2].height = 72
    build_help_sheet(workbook)

    PUBLIC_TEMPLATE_DIR.mkdir(parents=True, exist_ok=True)
    output_path = PUBLIC_TEMPLATE_DIR / "instruments-import-template.xlsx"
    return save_workbook(workbook, output_path)


if __name__ == "__main__":
    print(create_simplified_workbook())
    print(create_public_template())
