from __future__ import annotations

import json
import re
import time
import urllib.parse
import urllib.request
from difflib import SequenceMatcher
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
IMPORT_DIR = ROOT / "outputs" / "portal-import" / "wei-yuquan-2026-7"
PUBLICATION_FILE = IMPORT_DIR / "05-publications-wei-yuquan.xlsx"
CACHE_FILE = IMPORT_DIR / "publication_semantic_scholar_cache.json"
REPORT_FILE = IMPORT_DIR / "publication_semantic_scholar_report.md"


def normalize_title(value: str) -> str:
    text = value.lower()
    text = re.sub(r"[^a-z0-9\u4e00-\u9fff]+", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def similarity(a: str, b: str) -> float:
    left = normalize_title(a)
    right = normalize_title(b)
    if not left or not right:
        return 0.0
    if left == right:
        return 1.0
    if left in right or right in left:
        return 0.94
    return SequenceMatcher(None, left, right).ratio()


def fetch_json(url: str) -> dict:
    request = urllib.request.Request(
        url,
        headers={
            "User-Agent": "LabHub metadata enrichment (mailto:weiyq2019@cau.edu.cn)",
            "Accept": "application/json",
        },
    )
    with urllib.request.urlopen(request, timeout=20) as response:
        return json.loads(response.read().decode("utf-8"))


def lookup_by_doi(doi: str) -> dict:
    encoded = urllib.parse.quote(f"DOI:{doi}", safe="")
    fields = "title,abstract,year,venue,journal,externalIds"
    url = f"https://api.semanticscholar.org/graph/v1/paper/{encoded}?fields={fields}"
    data = fetch_json(url)
    return normalize_result(data, 1.0)


def lookup_by_title(title: str, year: str) -> dict:
    params = {
        "query": title,
        "limit": "5",
        "fields": "title,abstract,year,venue,journal,externalIds",
    }
    url = "https://api.semanticscholar.org/graph/v1/paper/search?" + urllib.parse.urlencode(params)
    data = fetch_json(url)
    best: dict | None = None
    best_score = 0.0
    for item in data.get("data", []):
        score = similarity(title, item.get("title") or "")
        item_year = str(item.get("year") or "")
        if year and item_year and item_year != year:
            score -= 0.08
        if score > best_score:
            best = item
            best_score = score
    if not best or best_score < 0.72:
        return {"matched": False, "score": round(best_score, 3)}
    return normalize_result(best, best_score)


def normalize_result(data: dict, score: float) -> dict:
    external = data.get("externalIds") or {}
    journal = data.get("journal") or {}
    return {
        "matched": True,
        "score": round(score, 3),
        "source": "Semantic Scholar",
        "title": data.get("title") or "",
        "doi": external.get("DOI") or "",
        "journal": journal.get("name") or data.get("venue") or "",
        "year": str(data.get("year") or ""),
        "volume": journal.get("volume") or "",
        "pages": journal.get("pages") or "",
        "abstract": (data.get("abstract") or "").strip(),
    }


def main() -> None:
    cache = json.loads(CACHE_FILE.read_text(encoding="utf-8")) if CACHE_FILE.exists() else {}
    workbook = load_workbook(PUBLICATION_FILE)
    sheet = workbook["Data"]
    headers = [cell.value for cell in sheet[1]]
    columns = {header: index + 1 for index, header in enumerate(headers)}

    attempted = 0
    matched = 0
    added_abstracts = 0
    added_dois = 0
    errors: list[str] = []

    for row_index in range(2, sheet.max_row + 1):
        title = str(sheet.cell(row_index, columns["title"]).value or "").strip()
        year = str(sheet.cell(row_index, columns["year"]).value or "").strip()
        doi = str(sheet.cell(row_index, columns["doi"]).value or "").strip()
        abstract = str(sheet.cell(row_index, columns["abstract"]).value or "").strip()
        if not title or abstract:
            continue

        key = f"{doi or title}::{year}"
        if key not in cache:
            attempted += 1
            try:
                cache[key] = lookup_by_doi(doi) if doi else lookup_by_title(title, year)
            except Exception as exc:
                cache[key] = {"matched": False, "error": str(exc)}
                errors.append(f"{title}: {exc}")
            CACHE_FILE.write_text(json.dumps(cache, ensure_ascii=False, indent=2), encoding="utf-8")
            time.sleep(1.05)

        item = cache[key]
        if item.get("matched"):
            matched += 1
        new_doi = str(item.get("doi") or "").strip()
        if new_doi and not doi:
            sheet.cell(row_index, columns["doi"]).value = new_doi
            added_dois += 1
        new_abstract = str(item.get("abstract") or "").strip()
        if len(new_abstract) >= 40:
            sheet.cell(row_index, columns["abstract"]).value = new_abstract
            added_abstracts += 1
        for field in ["journal", "volume", "pages"]:
            value = str(item.get(field) or "").strip()
            column = columns.get(field)
            current = str(sheet.cell(row_index, column).value or "").strip() if column else ""
            if column and value and not current:
                sheet.cell(row_index, column).value = value

    workbook.save(PUBLICATION_FILE)
    REPORT_FILE.write_text(
        "\n".join(
            [
                "# Semantic Scholar 补全报告",
                "",
                f"- 本轮新查询：{attempted}",
                f"- 命中记录：{matched}",
                f"- 新增 DOI：{added_dois}",
                f"- 新增摘要：{added_abstracts}",
                f"- 错误数：{len(errors)}",
                "",
                "## 错误样例",
                *(f"- {line}" for line in errors[:30]),
            ]
        ),
        encoding="utf-8",
    )
    print({"attempted": attempted, "matched": matched, "added_dois": added_dois, "added_abstracts": added_abstracts, "errors": len(errors)})


if __name__ == "__main__":
    main()
