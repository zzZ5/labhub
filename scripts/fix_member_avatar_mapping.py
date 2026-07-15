from __future__ import annotations

from io import BytesIO
from pathlib import Path

from docx import Document
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as ExcelImage
from PIL import Image, ImageOps


ROOT = Path(__file__).resolve().parents[1]
WORKBOOK_PATH = ROOT / "outputs" / "portal-import" / "wei-yuquan-2026-7" / "01-团队成员-魏雨泉-含毕业生.xlsx"
SOURCE_PATH = Path(r"C:/Users/baoju/Desktop/导入数据/魏雨泉简历 (2026.7).docx")
REFERENCE_PATH = ROOT / "outputs" / "_avatar-audit-20260715" / "original.xlsx"
PHOTO_COLUMN = 5


def cell_image(cell, document):
    blips = cell._tc.xpath(".//a:blip")
    if not blips:
        return None
    relation_id = blips[0].get("{http://schemas.openxmlformats.org/officeDocument/2006/relationships}embed")
    if not relation_id:
        return None
    part = document.part.related_parts.get(relation_id)
    if not part:
        return None
    rotation = 0.0
    node = blips[0]
    while node is not None:
        for descendant in node.iter():
            if descendant.tag.endswith("}xfrm") and descendant.get("rot"):
                rotation = float(descendant.get("rot")) / 60000
                break
        if rotation:
            break
        node = node.getparent()
    return part.blob, rotation


def source_photos(document):
    photos = {}
    # Current students take precedence when a name also appears in a graduate table.
    for table_index in (2, 0, 1):
        for row in document.tables[table_index].rows[2:]:
            name = " ".join(row.cells[1].text.split())
            if name and name not in photos:
                photos[name] = cell_image(row.cells[-1], document)
    return photos


def existing_images_by_row(sheet):
    images = {}
    for image in sheet._images:
        marker = getattr(image.anchor, "_from", None)
        if marker is not None:
            images[marker.row + 1] = image._data()
    return images


def workbook_photos_by_name(path):
    workbook = load_workbook(path, data_only=True)
    sheet = workbook.worksheets[1]
    images = existing_images_by_row(sheet)
    return {
        str(sheet.cell(row_number, 1).value or "").strip(): (image_data, 0.0)
        for row_number, image_data in images.items()
        if sheet.cell(row_number, 1).value
    }


def portrait_bytes(source, rotation=0.0):
    with Image.open(BytesIO(source)) as raw:
        image = ImageOps.exif_transpose(raw).convert("RGB")
        if rotation:
            image = image.rotate(-rotation, expand=True)
        target_ratio = 3 / 4
        width, height = image.size
        if width / height > target_ratio:
            crop_width = round(height * target_ratio)
            left = max(0, (width - crop_width) // 2)
            image = image.crop((left, 0, left + crop_width, height))
        elif width / height < target_ratio:
            crop_height = round(width / target_ratio)
            top = max(0, (height - crop_height) // 2)
            image = image.crop((0, top, width, top + crop_height))
        output = BytesIO()
        image.save(output, "JPEG", quality=94, subsampling=0, optimize=True)
        return output.getvalue()


def main():
    if not SOURCE_PATH.exists() or not WORKBOOK_PATH.exists():
        raise FileNotFoundError("缺少原始简历或成员导入文件。")
    document = Document(SOURCE_PATH)
    photos = workbook_photos_by_name(REFERENCE_PATH) if REFERENCE_PATH.exists() else source_photos(document)
    workbook = load_workbook(WORKBOOK_PATH)
    sheet = workbook.worksheets[1]
    existing = existing_images_by_row(sheet)
    lead_photo = photos.get("魏雨泉") or ((existing.get(2), 0.0) if existing.get(2) else None)
    sheet._images = []

    embedded = cleared = 0
    for row_number in range(2, sheet.max_row + 1):
        name = str(sheet.cell(row_number, 1).value or "").strip()
        photo = lead_photo if name == "魏雨泉" and lead_photo else photos.get(name)
        avatar_cell = sheet.cell(row_number, PHOTO_COLUMN)
        if not photo:
            avatar_cell.value = "清空"
            cleared += 1
            continue
        data = portrait_bytes(photo[0], photo[1])
        excel_image = ExcelImage(BytesIO(data))
        excel_image.width = 72
        excel_image.height = 96
        sheet.add_image(excel_image, avatar_cell.coordinate)
        avatar_cell.value = "已嵌入"
        sheet.row_dimensions[row_number].height = 76
        embedded += 1

    workbook.save(WORKBOOK_PATH)
    print(f"已更新 {WORKBOOK_PATH}")
    print(f"嵌入头像 {embedded} 张，无头像 {cleared} 人。")


if __name__ == "__main__":
    main()
