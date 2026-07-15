from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.datavalidation import DataValidation


TEMPLATE_DIR = Path("frontend/public/templates")

HEADER_FILL = PatternFill("solid", fgColor="EAF5EE")
HEADER_FONT = Font(bold=True, color="1F3D2B")
THIN_SIDE = Side(style="thin", color="E5E7EB")
CELL_BORDER = Border(left=THIN_SIDE, right=THIN_SIDE, top=THIN_SIDE, bottom=THIN_SIDE)


TEMPLATES = {
    "members-import-template.xlsx": {
        "title": "团队成员批量导入模板",
        "sheet_name": "团队成员",
        "notes": [
            "第二个 sheet 的列与“门户内容 - 团队成员 - 新增成员”表单对应，请从第 2 行开始填写。",
            "必填：姓名。邮箱建议填写，系统会优先按邮箱更新已有成员；邮箱为空时按姓名匹配。",
            "头像列用于放图片：可在对应行插入一张头像图片，系统会读取该行第一张图片；头像列里的文字不会被导入。",
            "展示排序：0 表示不在公开网站展示；大于 0 时按数字从小到大展示。",
        ],
        "headers": ["姓名", "身份头衔", "邮箱", "研究方向", "头像", "简介", "展示排序"],
        "rows": [
            ["张三", "博士研究生", "zhangsan@example.com", "堆肥微生物生态", "可在本行插入头像图片", "主要关注有机废弃物腐殖化过程。", 10],
        ],
        "widths": [14, 22, 30, 30, 24, 46, 12],
    },
    "publications-import-template.xlsx": {
        "title": "论文成果批量导入模板",
        "sheet_name": "论文成果",
        "notes": [
            "第二个 sheet 的列与“门户内容 - 论文成果 - 新增论文”表单对应，请从第 2 行开始填写。",
            "只需要填写 GB/T 7714-2025 格式引文，系统会自动拆分作者、题目、期刊、年份、卷期页和 DOI。",
            "摘要、可见范围、PDF附件、首页排序为管理字段；没有就留空。",
            "首页排序：0 表示不在首页展示；大于 0 时按数字从小到大展示。",
            "PDF 附件列只作提醒；批量导入先导入文字信息，PDF 请在单条论文编辑页面上传。",
        ],
        "headers": ["GB/T 7714-2025格式引文", "摘要", "可见范围", "PDF附件", "首页排序"],
        "rows": [
            ["作者1, 作者2. 示例论文题目. Environmental Research, 2026, 270: 121036. DOI: 10.0000/example", "填写论文摘要。", "公开", "导入后在单条论文中上传 PDF", 10],
        ],
        "widths": [82, 60, 14, 28, 12],
    },
    "projects-import-template.xlsx": {
        "title": "科研项目批量导入模板",
        "sheet_name": "科研项目",
        "notes": [
            "第二个 sheet 的列与“门户内容 - 科研项目 - 新增项目”表单对应，请从第 2 行开始填写。",
            "必填：项目名称。系统优先按项目编号更新已有项目；项目编号为空时按项目名称匹配。",
            "经费为可选字段，不需要公开展示时可留空。",
            "首页排序：0 表示不在首页展示；大于 0 时按数字从小到大展示。",
            "日期格式建议使用 YYYY-MM-DD。",
        ],
        "headers": ["项目名称", "项目编号", "资助来源", "负责人", "状态", "开始日期", "结束日期", "经费", "可见范围", "说明", "首页排序"],
        "rows": [
            ["有机废弃物资源化与堆肥过程调控研究", "项目编号示例", "国家级/省部级科研项目", "团队负责人", "在研", "2026-01-01", "2028-12-31", "", "公开", "门户展示版说明。", 10],
        ],
        "widths": [42, 22, 30, 16, 12, 14, 14, 10, 14, 46, 12],
    },
    "patents-import-template.xlsx": {
        "title": "专利成果批量导入模板",
        "sheet_name": "专利成果",
        "notes": [
            "第二个 sheet 的列与“门户内容 - 专利成果 - 新增专利”表单对应，请从第 2 行开始填写。",
            "必填：专利名称。系统优先按专利号更新已有专利；专利号为空时按专利名称匹配。",
            "日期格式建议使用 YYYY-MM-DD。",
            "PDF 附件列只作提醒；批量导入先导入文字信息，PDF 请在单条专利编辑页面上传。",
        ],
        "headers": ["专利名称", "专利号", "状态", "发明人", "PDF附件", "申请日期", "授权日期", "可见范围", "首页排序"],
        "rows": [
            ["一种堆肥过程调控相关专利", "CN000000000A", "已申请", "发明人1; 发明人2", "导入后在单条专利中上传 PDF", "2026-01-01", "", "公开", 10],
        ],
        "widths": [42, 24, 14, 34, 28, 14, 14, 14, 12],
    },
    "awards-import-template.xlsx": {
        "title": "获奖成果批量导入模板",
        "sheet_name": "获奖成果",
        "notes": [
            "第二个 sheet 的列与“门户内容 - 获奖成果 - 新增获奖”表单对应，请从第 2 行开始填写。",
            "必填：奖项名称。系统按奖项名称和获奖日期更新已有获奖成果。",
            "获奖图片列用于放图片：可在对应行插入一张获奖图片，系统会读取该行第一张图片；获奖图片列里的文字不会被导入。",
            "附件 PDF 列只作提醒；批量导入先导入文字信息，PDF 或其它附件请在单条获奖成果编辑页面上传。",
        ],
        "headers": ["奖项名称", "奖项等级", "获奖日期", "参与人员", "获奖图片", "附件PDF", "可见范围", "说明", "首页排序"],
        "rows": [
            ["示例获奖成果", "一等奖", "2026-01-01", "团队成员", "可在本行插入获奖图片", "导入后在单条获奖中上传附件", "公开", "填写获奖说明。", 10],
        ],
        "widths": [38, 20, 14, 32, 24, 28, 14, 44, 12],
    },
}


def add_validation(sheet, headers: list[str], header: str, choices: list[str]) -> None:
    if header not in headers:
        return
    column = headers.index(header) + 1
    column_letter = sheet.cell(row=1, column=column).column_letter
    validation = DataValidation(type="list", formula1=f'"{",".join(choices)}"', allow_blank=True)
    sheet.add_data_validation(validation)
    validation.add(f"{column_letter}2:{column_letter}500")


def write_template(filename: str, config: dict) -> None:
    workbook = Workbook()
    info_sheet = workbook.active
    info_sheet.title = "导入说明"
    info_sheet["A1"] = config["title"]
    info_sheet["A1"].font = Font(bold=True, size=16, color="1F3D2B")
    info_sheet["A2"] = "填写说明"
    info_sheet["A2"].font = Font(bold=True, color="00873C")
    for index, note in enumerate(config["notes"], start=3):
        cell = info_sheet[f"A{index}"]
        cell.value = f"{index - 2}. {note}"
        cell.alignment = Alignment(wrap_text=True, vertical="top")
    info_sheet.column_dimensions["A"].width = 105

    data_sheet = workbook.create_sheet(config["sheet_name"])
    headers = config["headers"]
    data_sheet.append(headers)
    for row in config["rows"]:
        data_sheet.append(row)

    for column, width in enumerate(config["widths"], start=1):
        data_sheet.column_dimensions[data_sheet.cell(row=1, column=column).column_letter].width = width
    for cell in data_sheet[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = CELL_BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for row in data_sheet.iter_rows(min_row=2, max_row=max(2, data_sheet.max_row), max_col=len(headers)):
        for cell in row:
            cell.border = CELL_BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    data_sheet.freeze_panes = "A2"
    add_validation(data_sheet, headers, "可见范围", ["公开", "成员可见", "管理员可见"])
    workbook.save(TEMPLATE_DIR / filename)


def main() -> None:
    TEMPLATE_DIR.mkdir(parents=True, exist_ok=True)
    for filename, config in TEMPLATES.items():
        write_template(filename, config)
        print(TEMPLATE_DIR / filename)


if __name__ == "__main__":
    main()
